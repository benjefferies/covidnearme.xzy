import io
import json
import os
import collections

import googlemaps
import requests
import boto3
from datetime import datetime, timedelta

from botocore.exceptions import ClientError
from dateutil.parser import parse
from googlemaps.places import find_place, place

from openpyxl import load_workbook

bucket = "covidnearme.xyz"
cases_key = "coronavirus-data.json"
districts_key = "districts.json"
hospital_lookup_key = "hospital-lookup.json"


def handler(event, context):
    s3 = boto3.client('s3')
    district_cases = get_cases()
    deaths = get_deaths()
    hospital_lookup = get_hospital_to_district(s3, deaths)
    district_to_deaths = map_district_to_deaths(hospital_lookup, deaths)
    merged_data = merge(district_cases, district_to_deaths)
    fill_empties(merged_data)
    upload_cases(merged_data, s3)


def get_hospital_to_district(s3, deaths):
    gmaps = googlemaps.Client(key=os.getenv("GOOGLE_PLACES_API_KEY"))
    hospital_lookup = {}
    try:
        hospital_lookup_obj = s3.get_object(Bucket=bucket, Key=hospital_lookup_key)
        hospital_lookup = json.loads(hospital_lookup_obj['Body'].read())
    except ClientError as e:
        if e.response['Error']['Code'] == "404":
            print(f"File {hospital_lookup_key} does not exist yet")
    new_hospitals = deaths.keys() - hospital_lookup.keys()
    print(f"Need to look up hospitals {new_hospitals}")
    for hospital in new_hospitals:
        print(f"Looking up hospital {hospital}")
        places = find_place(gmaps, [hospital], "textquery", fields=["place_id"])
        if len(places["candidates"]) == 0:
            print(f"Warning could not find address for {hospital}. Skipping...")
            continue
        place_details = place(gmaps, places["candidates"][0]['place_id'])
        district = next(iter([p["long_name"] for p in place_details["result"]["address_components"] if "administrative_area_level_2" in p["types"]]))
        if len(places) > 1:
            print(f"Using {district} more than one address for {hospital}, using data {', '.join([p['short_name'] for p in place_details['result']['address_components']])}")
        hospital_lookup[hospital] = district
    print("Uploading hospital data")
    s3.put_object(Bucket=bucket, Key=hospital_lookup_key, Body=bytes(json.dumps(hospital_lookup), "UTF-8"))
    return hospital_lookup


def fill_empties(data):
    for district in data.keys():
        for date in data[district]:
            if "casesDaily" not in data[district][date]: data[district][date]["casesDaily"] = 0
            if "deathsDaily" not in data[district][date]: data[district][date]["deathsDaily"] = 0


def merge(d, u):
    for k, v in u.items():
        if isinstance(v, collections.abc.Mapping):
            d[k] = merge(d.get(k, {}), v)
        else:
            d[k] = v
    return d


def map_district_to_deaths(hospital_lookup, deaths):
    districts_deaths_per_day = {}
    for hospital in deaths.keys():
        district = hospital_lookup[hospital]
        if district in districts_deaths_per_day:
            districts_deaths_per_day[district] = {date: {"deathsDaily": districts_deaths_per_day[district][date]["deathsDaily"] or 0 + deaths[hospital][date]} or 0 for date in deaths[hospital].keys()}
        else:
            districts_deaths_per_day[district] = {date: {"deathsDaily": deaths[hospital][date]} for date in deaths[hospital].keys()}
    calculate_death_totals(districts_deaths_per_day)

    if "City of Bristol" in districts_deaths_per_day:
        districts_deaths_per_day["Bristol, City of"] = districts_deaths_per_day["City of Bristol"]
        del districts_deaths_per_day["City of Bristol"]
    return districts_deaths_per_day


def calculate_death_totals(districts_deaths_per_day):
    for district, dates in districts_deaths_per_day.items():
        print(f"Adding totals for {district}")
        previous = {"deathsTotal": 0}
        for date in sorted(dates):
            districts_deaths_per_day[district][date]["deathsTotal"] = previous["deathsTotal"] + districts_deaths_per_day[district][date]["deathsDaily"]
            previous = districts_deaths_per_day[district][date]


def get_deaths():
    now = datetime.now()
    file_uri = get_deaths_url(now)
    print(f"Getting data from {file_uri}")
    response = requests.get(file_uri)
    if response.status_code != 200:
        file_uri = get_deaths_url(now - timedelta(days=1))
        print(f"Today not found. Getting yesterdays data from {file_uri}")
        response = requests.get(file_uri)
    if response.status_code != 200:
        print("Could not find data")
        exit(1)
    content = io.BytesIO(response.content)
    wb = load_workbook(content)
    trust_sheet_name = next(iter([sheet for sheet in wb.sheetnames if "by trust" in sheet]))
    print(f"Found sheet with deaths by trust {trust_sheet_name}")
    trust_sheet = wb[trust_sheet_name]
    start_of_table = next(iter([cell.row for cell in trust_sheet["b"] if "nhs england region" in str(cell.value).lower()]))
    row_numbers = [cell.row for cell in trust_sheet["b"][start_of_table:] if "-" not in str(cell.value) and cell.value is not None]
    date_columns = [{"col": date[start_of_table - 1].column_letter, "date": date[start_of_table - 1].value} for date in trust_sheet.iter_cols() if is_date(str(date[start_of_table - 1].value))]
    trust_data = {}
    for row_number in row_numbers:
        hospital = trust_sheet["e"][row_number].value
        trust_data[hospital] = {col["date"].isoformat(): trust_sheet[col["col"]][row_number].value for col in date_columns}
    if None in trust_data:
        del trust_data[None]
    return trust_data


def get_deaths_url(date):
    date_path = f"2/{date.strftime('%Y')}/{date.strftime('%m')}"  # Date path is zero offset
    date_file = f"{int(date.strftime('%d'))}-{date.strftime('%b-%Y')}"
    file_uri = f"https://www.england.nhs.uk/statistics/wp-content/uploads/sites/{date_path}/COVID-19-total-announced-deaths-{date_file}.xlsx"
    return file_uri


def upload_cases(data, s3):
    print("Extracting districts")
    print("Uploading cases")
    s3.put_object(Bucket=bucket, Key=cases_key, Body=bytes(json.dumps(data), "UTF-8"))
    print("Uploading districts")
    s3.put_object(Bucket=bucket, Key=districts_key, Body=bytes(json.dumps(list(data.keys())), "UTF-8"))
    print("Completed")


def get_cases():
    print("Downloading latest cases")
    response = requests.get("https://c19downloads.azureedge.net/downloads/json/coronavirus-cases_latest.json")
    covid_data = response.json()
    district_cases = collections.defaultdict(dict)
    for case in covid_data["utlas"]:
        date = parse(case["specimenDate"], fuzzy=False)
        district_cases[case["areaName"]][date.isoformat()] = {
            "casesDaily": case["dailyLabConfirmedCases"],
            "casesTotal": case["totalLabConfirmedCases"],
        }
    return district_cases


def is_date(string, fuzzy=False):
    """
    Return whether the string can be interpreted as a date.

    :param string: str, string to check for date
    :param fuzzy: bool, ignore unknown tokens in string if True
    """
    try:
        parse(string, fuzzy=fuzzy)
        return True

    except ValueError:
        return False


if __name__ == '__main__':
    handler({}, {})
